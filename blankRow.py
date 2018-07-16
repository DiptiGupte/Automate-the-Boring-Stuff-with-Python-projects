import openpyxl

#tkaes two integers,N and M, and a filename string
#starting at row N, should insert M blank rows into spreadsheet
def blankRowInserter(N, M, spreadsheet):
    newWb = openpyxl.Workbook()
    newSheet = wb.active
    givenWb = openpyxl.load_workbook(spreadsheet)
    givenSheet = read.active
    for rowNum in range(1, givenSheet.max_row + 1):
        for colNum in range(1, givenSheet.max_col + 1):
            if rowNum < M:
                newSheet.cell(row = rowNum, column = colNum).value = givenSheet.cell(row = rowNum, column = colNum).value
            else:
                newSheet.cell(row = rowNum + M, column = colNum).value = givenSheet.cell(row = rowNum, column = colNum).value
    newname = 'copy_' + spreadsheet
    newSheet.save(newname)

blankRowInserter(3, 2, 'myProduce.xlsx')
