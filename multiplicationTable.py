import openpyxl
from openpyxl.styles import Font

#takes a number N from command line and creates an NxN multiplication table in an excel spreadsheet
def multiTable(N):
    wb = openpyxl.Workbook()
    sheet = wb.active
    boldFont = Font(bold = True)
    #entering number of rows
    print('rows')
    for rowNum in range(2, N+2):  #skip the first row
        sheet.cell(row = rowNum, column = 1).value = rowNum - 1
        sheet.font = boldFont
        print(rowNum - 1)
    print('col')
    for colNum in range(2, N+2):    #skip the first column
        sheet.cell(row = 1, column = colNum).value = colNum - 1
        sheet.font = boldFont
    #filling in table
    for r in range(2, N+2):
        for c in range(2, N+2):
            sheet.cell(row = r, column = c).value = (r - 1) * (c - 1)
    wb.save('NxN.xlsx')

def main():
    print('Enter a number: ')
    num = input()
    multiTable(num)
    print('done')

main()
