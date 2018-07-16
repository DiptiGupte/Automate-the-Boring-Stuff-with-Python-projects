import openpyxl

#invert the row and colum of the cells in spreadsheet
def invert(spreadsheet):
    newWb = openpyxl.Workbook()
    newSheet = wb.active
    givenWb = openpyxl.load_workbook(spreadsheet)
    givenSheet = read.active
    for rowNum in range(1, givenSheet.max_row + 1):
        for colNum in range(1, givenSheet.max_col + 1):
            newSheet.cell(row = colNum, column = rowNum).value = givenSheet.cell(row = rowNum, column = colNum).value
    newname = 'inverted_' + spreadsheet
    givenSheet.save(newname)

invert('myProduce.xlsx')
